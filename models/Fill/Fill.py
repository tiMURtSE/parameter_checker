from openpyxl.styles import PatternFill
from openpyxl.cell.cell import Cell
from models.Product.Product import Product
from typing import List

class Fill:
    RED_COLOR = "F73434"
    ORANGE_COLOR = "FC9003"

    def __init__(self):
        pass

    def fill_cells(self, product: Product):
        required_params = product.required_params

        for param in required_params:
            if param.is_required_parameter_unfilled():
                if param.is_conditional:
                    self._fill_cell_orange(param.cell)
                else:
                    self._fill_cell_red(param.cell)

    def _fill_cell_red(self, cell: Cell):
        red_fill = PatternFill(start_color=self.RED_COLOR, end_color=self.RED_COLOR, fill_type="solid")

        cell.fill = red_fill

    def _fill_cell_orange(self, cell: Cell):
        orange_fill = PatternFill(start_color=self.ORANGE_COLOR, end_color=self.ORANGE_COLOR, fill_type="solid")

        cell.fill = orange_fill