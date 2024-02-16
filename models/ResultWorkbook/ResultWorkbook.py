import sys
import datetime
import openpyxl
import tkinter as tk
from tkinter import filedialog

from local_types import ResultType

class ResultWorkbook:
    def __init__(self):
        self._workbook = None
        self._sheet = None
        self._directory_path = None

    def write_result(self, result: ResultType):
        self._directory_path = self._input_result_directory_path()
        self._create_workbook()

        self._write_total_products_info(result)
        self._write_headings()
        self._write_params_stats(result)

        self._save_workbook()

    def _create_workbook(self):
        self._workbook = openpyxl.Workbook()
        self._sheet = self._workbook.active

    def _input_result_directory_path(self):
        print("\nВыберите папку для вывода результатов: ")
        self._remove_extra_window()

        directory_path = filedialog.askdirectory()

        if not directory_path:
            print("Папка не была выбрана")
            sys.exit()

        return directory_path
    
    def _remove_extra_window(self):
        root = tk.Tk()
        root.withdraw()

    def _save_workbook(self):
        filename = self._create_filename()
        self._workbook.save(filename=filename)

    def _write_total_products_info(self, result: ResultType):
        total_products = result["total_products"]
        total_incomplete_products = result["total_incomplete_products"]

        self._sheet["A1"].value = f"У {total_incomplete_products} из {total_products} товаров не заполнены обязательные параметры"

    def _create_filename(self):
        current_datetime = self._get_current_datetime()
        filename = f"{self._directory_path}/Статистика_по_незаполненным_параметрам_{current_datetime}.xlsx"

        return filename

    def _get_current_datetime(self):
        current_datetime = datetime.datetime.now()
        formatted_datetime = current_datetime.strftime("%d-%m-%Y_%H-%M-%S")

        return formatted_datetime
    
    def _write_headings(self):
        param_name_heading = "Название параметра"
        total_unfilled_params_heading = "Количество товаров c незаполненным параметром"

        self._sheet["A3"].value = param_name_heading
        self._sheet["B3"].value = total_unfilled_params_heading

    def _write_params_stats(self, product: ResultType):
        for index, (param_name, products_with_unfiiled_param) in enumerate(product["unfilled_required_params"].items()):
            self._sheet.cell(row=index + 4, column=1).value = param_name
            self._sheet.cell(row=index + 4, column=2).value = products_with_unfiiled_param